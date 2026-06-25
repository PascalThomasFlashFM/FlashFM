"""
Logique de traitement : lecture du fichier mensuel RRR et mise à jour du fichier de synthèse.
"""

from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from typing import Callable, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

LogFn = Callable[[str, str], None]

# Correspondance mot-clé zone mensuelle → libellé section synthèse
ZONE_MAP = {
    "brive": "BRIVE",
    "gueret": "GUERET",
    "guéret": "GUERET",
    "limoges": "LIMOGES",
    "montmorillon": "MONTMORILLON",
}

MONTHS_FR = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre",
]

ALL_SECTIONS = {"LIMOGES", "GUERET", "MONTMORILLON", "BRIVE", "FWD"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalize(s) -> str:
    """Normalise une chaîne pour la comparaison (minuscules, sans accents, sans espaces doubles)."""
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.lower().split())


def _zone_key(zone_label: str) -> Optional[str]:
    n = _normalize(zone_label)
    for keyword, section in ZONE_MAP.items():
        if keyword in n:
            return section
    return None


def _month_col(month: int) -> tuple[int, int]:
    """Retourne les numéros de colonnes (1-indexed) Brut et Net pour le mois (1-12)."""
    brut_col = 2 + (month - 1) * 2   # Janvier→B(2), Février→D(4), Mars→F(6)…
    return brut_col, brut_col + 1


def _safe_str(row, idx: int) -> str:
    if len(row) > idx and row[idx] is not None:
        return str(row[idx]).strip()
    return ""


def _safe_float(row, idx: int) -> float:
    if len(row) > idx and isinstance(row[idx], (int, float)):
        return float(row[idx])
    return 0.0


# ---------------------------------------------------------------------------
# Lecture du fichier mensuel
# ---------------------------------------------------------------------------

class ZoneData:
    def __init__(self, name: str):
        self.name = name
        self.commerciaux: dict[str, dict] = defaultdict(lambda: {"brut": 0.0, "net": 0.0})
        self.production: float = 0.0

    def add_line(self, commercial: str, brut: float, net: float):
        self.commerciaux[commercial]["brut"] += brut
        self.commerciaux[commercial]["net"] += net


def _detect_year_month(ws) -> tuple[int, int]:
    """Détecte année et mois dans l'onglet Régie."""
    year = month = None
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, (int, float)):
                v = int(cell)
                if 2000 <= v <= 2100 and year is None:
                    year = v
                elif 1 <= v <= 12 and month is None and year is not None:
                    month = v
        if year and month:
            break
    if not year or not month:
        raise ValueError("Impossible de détecter l'année et le mois dans l'onglet Régie.")
    return year, month


def parse_monthly_file(path: str, log: LogFn) -> tuple[int, int, dict[str, ZoneData], ZoneData]:
    """Lit le fichier mensuel et retourne (année, mois, zones, fwd_data)."""
    wb = load_workbook(path, data_only=True)

    ws_regie = wb["Régie"]
    year, month = _detect_year_month(ws_regie)
    log(f"  Période détectée : {MONTHS_FR[month-1]} {year}", "info")

    zones: dict[str, ZoneData] = {}

    for row in ws_regie.iter_rows(values_only=True):
        year_val = row[1] if len(row) > 1 else None
        zone_label = _safe_str(row, 5)   # col F
        commercial = _safe_str(row, 4)   # col E
        brut = _safe_float(row, 13)      # col N
        net = _safe_float(row, 15)       # col P
        production = _safe_float(row, 16)  # col Q

        section = _zone_key(zone_label)
        if not section:
            continue

        if section not in zones:
            zones[section] = ZoneData(section)

        # Ligne de total de zone : année absente, commercial absent, net > 0
        if not isinstance(year_val, (int, float)) and not commercial and net > 0:
            zones[section].production = production
            continue

        if isinstance(year_val, (int, float)) and commercial:
            zones[section].add_line(commercial, brut, net)

    for section, zd in zones.items():
        log(f"  Zone {section} : {len(zd.commerciaux)} commercial(aux)", "")
        for comm, v in zd.commerciaux.items():
            log(f"    {comm} → Brut {v['brut']:.2f} € / Net {v['net']:.2f} €", "")
        if zd.production:
            log(f"    Production : {zd.production:.2f} €", "")

    # FWD
    fwd_data = ZoneData("FWD")
    if "FWD" in wb.sheetnames:
        ws_fwd = wb["FWD"]
        for row in ws_fwd.iter_rows(values_only=True):
            if not isinstance(row[1] if len(row) > 1 else None, (int, float)):
                continue
            commercial = _safe_str(row, 4)
            brut = _safe_float(row, 13)
            net = _safe_float(row, 15)
            if commercial:
                fwd_data.add_line(commercial, brut, net)

        log(f"  FWD : {len(fwd_data.commerciaux)} commercial(aux)", "")
        for comm, v in fwd_data.commerciaux.items():
            log(f"    {comm} → Brut {v['brut']:.2f} € / Net {v['net']:.2f} €", "")

    return year, month, zones, fwd_data


# ---------------------------------------------------------------------------
# Structure du fichier de synthèse
# ---------------------------------------------------------------------------

class SectionInfo:
    """Localisation dynamique d'une section dans la feuille de synthèse."""

    def __init__(self, header_row: int, data_start: int, sum_row: int, production_row: Optional[int]):
        self.header_row = header_row
        self.data_start = data_start
        self.sum_row = sum_row          # ligne avec =SUM(…) – les données s'arrêtent avant
        self.production_row = production_row

    @property
    def data_end(self) -> int:
        """Dernière ligne de donnée = juste avant sum_row."""
        return self.sum_row - 1

    def shift(self, from_row: int, by: int = 1):
        """Décale tous les indices >= from_row suite à une insertion de ligne."""
        if self.data_start >= from_row:
            self.data_start += by
        if self.sum_row >= from_row:
            self.sum_row += by
        if self.production_row and self.production_row >= from_row:
            self.production_row += by


def _is_sum_formula(val) -> bool:
    if isinstance(val, str):
        return val.strip().upper().startswith("=SUM(")
    return False


def _is_summary_label(val) -> bool:
    if not isinstance(val, str):
        return False
    v = _normalize(val)
    return (
        v.startswith("ca total")
        or v.startswith("total")
        or "production" in v
        or "synthese" in v
    )


def find_sections(ws) -> dict[str, SectionInfo]:
    """
    Détecte les sections en parcourant col A.
    Fonctionne avec data_only=False pour voir les formules =SUM.
    """
    sections: dict[str, SectionInfo] = {}

    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        val = cell.value
        if not isinstance(val, str):
            continue
        key = val.strip().upper()
        if key not in ALL_SECTIONS:
            continue
        # Ne garder que la première occurrence de chaque section
        if key in sections:
            continue

        header_row = cell.row
        # Structure : header_row → noms mois / Brut-Net ; header_row+1 → Brut/Net ; header_row+2 → 1ère donnée
        # (vérifiée sur l'exemple fourni)
        data_start = header_row + 2

        # Chercher la ligne SUM : scan vers le bas
        sum_row = None
        production_row = None

        for r in range(data_start, data_start + 80):
            if r > ws.max_row:
                break
            a_val = ws.cell(r, 1).value
            b_val = ws.cell(r, 2).value

            # Ligne SUM : col B est une formule =SUM
            if _is_sum_formula(str(b_val) if b_val is not None else ""):
                sum_row = r
                # Chercher la ligne production juste après
                for pr in range(r + 1, r + 5):
                    if pr > ws.max_row:
                        break
                    pv = ws.cell(pr, 1).value
                    if isinstance(pv, str) and ("roduction" in pv or "roduion" in pv or "rodut" in pv):
                        production_row = pr
                        break
                break

            # Fin de section si on tombe sur un autre header de section
            if isinstance(a_val, str) and a_val.strip().upper() in ALL_SECTIONS and r != header_row:
                # Pas de SUM trouvée, prendre la ligne précédente
                sum_row = r - 1
                break

        if sum_row is None:
            # Fallback : utiliser data_start + 30
            sum_row = data_start + 30

        sections[key] = SectionInfo(header_row, data_start, sum_row, production_row)

    return sections


# ---------------------------------------------------------------------------
# Résolution des noms dans la section FWD (formules =Ax)
# ---------------------------------------------------------------------------

def _resolve_name(ws, row: int) -> Optional[str]:
    """Lit le nom du commercial en cellule A{row}, en résolvant les =A{n}."""
    val = ws.cell(row, 1).value
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        m = re.match(r"^=A(\d+)$", val, re.IGNORECASE)
        if m:
            ref = ws.cell(int(m.group(1)), 1).value
            return str(ref).strip() if ref else None
        if val.startswith("="):
            return None
        return val or None
    return str(val).strip() or None


# ---------------------------------------------------------------------------
# Recherche / création de ligne commerciale
# ---------------------------------------------------------------------------

def _insert_rows_fix_merges(ws, insert_at: int):
    """
    Insère une ligne à insert_at et met à jour toutes les plages de cellules fusionnées.
    openpyxl ne décale pas les merged_cells lors de insert_rows(), ce qui corromprait
    les cellules situées sous une fusion non déplacée.
    """
    ws.insert_rows(insert_at)

    # Collecter et recalculer toutes les plages fusionnées
    old_ranges = list(ws.merged_cells.ranges)
    # Vider la liste (on va tout réajouter recalculé)
    for mr in old_ranges:
        ws.merged_cells.ranges.discard(mr)

    for mr in old_ranges:
        min_r = mr.min_row
        max_r = mr.max_row
        min_c = mr.min_col
        max_c = mr.max_col
        if min_r >= insert_at:
            min_r += 1
            max_r += 1
        elif max_r >= insert_at:
            max_r += 1
        ws.merge_cells(
            start_row=min_r, start_column=min_c,
            end_row=max_r, end_column=max_c,
        )


def find_or_create_commercial_row(
    ws,
    section: SectionInfo,
    all_sections: dict[str, SectionInfo],
    commercial_name: str,
    log: LogFn,
) -> tuple[int, bool]:
    """
    Retourne (numéro de ligne, was_inserted).
    Crée une nouvelle ligne si nécessaire et met à jour toutes les SectionInfo.
    """
    target_norm = _normalize(commercial_name)

    # Chercher le commercial dans les lignes existantes
    for r in range(section.data_start, section.sum_row):
        name = _resolve_name(ws, r)
        if name and _normalize(name) == target_norm:
            return r, False

    # Pas trouvé → chercher une cellule vide dans la plage
    for r in range(section.data_start, section.sum_row):
        if _resolve_name(ws, r) is None:
            ws.cell(r, 1).value = commercial_name
            log(f"    + '{commercial_name}' → ligne {r} (cellule vide)", "warn")
            return r, False

    # Aucune place : insérer une ligne juste avant sum_row
    insert_at = section.sum_row
    _insert_rows_fix_merges(ws, insert_at)
    ws.cell(insert_at, 1).value = commercial_name

    # Mettre à jour TOUTES les SectionInfo dont les indices sont >= insert_at
    for si in all_sections.values():
        si.shift(insert_at)

    log(f"    + '{commercial_name}' → ligne {insert_at} (insertion)", "warn")
    return insert_at, True


# ---------------------------------------------------------------------------
# Mise à jour de la synthèse
# ---------------------------------------------------------------------------

def _fix_sum_formulas(ws, section: SectionInfo, log: LogFn, section_name: str):
    """
    Corrige les formules =SUM dans la ligne sum_row pour couvrir la plage data_start→sum_row-1.
    Nécessaire après insertion de lignes (openpyxl ne met pas à jour les formules automatiquement).
    """
    sum_row = section.sum_row
    data_start = section.data_start
    data_end = sum_row - 1

    if data_end < data_start:
        return

    updated = 0
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(sum_row, col)
        val = cell.value
        if isinstance(val, str) and val.strip().upper().startswith("=SUM("):
            col_letter = get_column_letter(col)
            cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
            updated += 1

    if updated:
        log(f"    Formules SUM {section_name} mises à jour (données : lignes {data_start}→{data_end})", "")


def update_synthesis(
    synthesis_path: str,
    year: int,
    month: int,
    zones: dict[str, ZoneData],
    fwd_data: ZoneData,
    log: LogFn,
):
    log(f"\n{'─'*60}", "")
    log("Mise à jour du fichier de synthèse :", "info")
    log(f"  {synthesis_path}", "")

    # Charger avec data_only=False pour voir les formules =SUM
    wb = load_workbook(synthesis_path)
    sheet_name = str(year)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"L'onglet '{sheet_name}' n'existe pas dans le fichier de synthèse.\n"
            "Veuillez créer cet onglet avant de lancer le traitement."
        )

    ws = wb[sheet_name]
    log(f"  Onglet '{sheet_name}' trouvé.", "ok")

    sections = find_sections(ws)
    log(f"  Sections détectées : {list(sections.keys())}", "")
    for name, si in sections.items():
        log(
            f"    {name} : header={si.header_row}, données={si.data_start}→{si.data_end}, "
            f"SUM={si.sum_row}, production={si.production_row}",
            "",
        )

    brut_col, net_col = _month_col(month)
    log(
        f"  Colonnes {MONTHS_FR[month-1]} : "
        f"Brut={get_column_letter(brut_col)}, Net={get_column_letter(net_col)}",
        "",
    )

    # Ordre de traitement : du bas vers le haut pour éviter les décalages
    all_items: list[tuple[str, ZoneData]] = []
    if "FWD" in sections and fwd_data.commerciaux:
        all_items.append(("FWD", fwd_data))
    for name, zd in zones.items():
        if name in sections:
            all_items.append((name, zd))

    # Trier du bas vers le haut (header_row décroissant)
    all_items.sort(key=lambda x: sections[x[0]].header_row, reverse=True)

    inserted_sections: set[str] = set()

    for section_name, data in all_items:
        sec = sections[section_name]
        log(f"\n  Section {section_name} :", "head")

        section_had_insert = False
        for commercial, vals in data.commerciaux.items():
            row, was_inserted = find_or_create_commercial_row(ws, sec, sections, commercial, log)
            if was_inserted:
                section_had_insert = True
            ws.cell(row, brut_col).value = round(vals["brut"], 2)
            ws.cell(row, net_col).value = round(vals["net"], 2)
            log(
                f"    {commercial} → ligne {row} : Brut={vals['brut']:.2f} €, Net={vals['net']:.2f} €",
                "ok",
            )

        if section_had_insert:
            inserted_sections.add(section_name)

        # Production (uniquement pour les zones Régie)
        if data.production and sec.production_row:
            ws.cell(sec.production_row, net_col).value = round(data.production, 2)
            log(f"    Production → ligne {sec.production_row} : {data.production:.2f} €", "ok")

    # Corriger les formules SUM dans toutes les sections qui ont eu des insertions
    # (et aussi dans les sections en-dessous qui ont été décalées)
    for section_name, sec in sections.items():
        _fix_sum_formulas(ws, sec, log, section_name)

    wb.save(synthesis_path)
    log(f"\n  Fichier sauvegardé : {synthesis_path}", "ok")


# ---------------------------------------------------------------------------
# Point d'entrée
# ---------------------------------------------------------------------------

def process_monthly_file(monthly_path: str, synthesis_path: str, log: LogFn):
    log("═" * 60, "head")
    log("  TRAITEMENT CA RRR - Flash FM", "head")
    log("═" * 60, "head")
    log(f"\nFichier mensuel : {monthly_path}", "info")
    log("Lecture et analyse en cours…", "")

    year, month, zones, fwd_data = parse_monthly_file(monthly_path, log)
    update_synthesis(synthesis_path, year, month, zones, fwd_data, log)
