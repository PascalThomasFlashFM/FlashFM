import unicodedata
from copy import copy
from dataclasses import dataclass

import openpyxl
from openpyxl.styles import Font

MONTH_COLUMNS = {
    1: "C", 2: "D", 3: "E", 4: "F", 5: "G", 6: "H",
    7: "I", 8: "J", 9: "K", 10: "L", 11: "M", 12: "N",
}
FIRST_DATA_ROW = 3
LAST_DATA_ROW = 173
RECETTE_LAST_ROW = 16  # rows 3-16 are receipts, rows after are expenses
UNMATCHED_SHEET = "À vérifier"
ESTIMATE_FONT_COLOR = "FFC5D9F1"  # light blue used for forecast/estimate values
REAL_FONT_COLOR = "FF000000"


def normalize(text: str) -> str:
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    return " ".join(text.upper().split())


@dataclass
class TargetRow:
    row: int
    bloc: str
    label: str
    is_recette: bool


def load_workbook(path: str):
    return openpyxl.load_workbook(path)


def get_treasury_sheet(wb, year: int):
    sheet_name = f"Trésorerie {year}"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Onglet '{sheet_name}' introuvable dans le classeur.")
    return wb[sheet_name]


MIN_ROLLOVER_YEAR = 2020  # safety floor: never auto-create sheets below this


def ensure_year_sheet(wb, year: int):
    """Returns the "Trésorerie {year}" sheet, creating it (and any missing
    year in between) by cloning the previous year's sheet if needed. Every
    detail cell in the new sheet is seeded with the previous year's value as
    a blue estimate — real Pennylane data will progressively replace it
    through the year, exactly like any other forecast cell."""
    sheet_name = f"Trésorerie {year}"
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    if year - 1 < MIN_ROLLOVER_YEAR:
        raise ValueError(
            f"Onglet '{sheet_name}' introuvable et aucun onglet antérieur "
            f"disponible pour servir de modèle."
        )

    prev_ws = ensure_year_sheet(wb, year - 1)
    new_ws = wb.copy_worksheet(prev_ws)
    new_ws.title = sheet_name

    targets = parse_target_rows(prev_ws)
    for month_col in MONTH_COLUMNS.values():
        for target in targets:
            prev_value = prev_ws[f"{month_col}{target.row}"].value
            new_cell = new_ws[f"{month_col}{target.row}"]
            if isinstance(prev_value, (int, float)):
                new_cell.value = prev_value
            _set_font_color(new_cell, ESTIMATE_FONT_COLOR)
    return new_ws


def parse_target_rows(ws) -> list[TargetRow]:
    """Scans column A (bloc, forward-filled) and column B (fournisseur/label)
    to build the list of rows that can receive an amount. Rows where column B
    is empty are totals/separators and are skipped."""
    targets = []
    current_bloc = None
    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        bloc_cell = ws.cell(row=row, column=1).value
        label_cell = ws.cell(row=row, column=2).value
        if bloc_cell:
            current_bloc = bloc_cell
        if label_cell:
            targets.append(
                TargetRow(
                    row=row,
                    bloc=current_bloc or "",
                    label=label_cell,
                    is_recette=row <= RECETTE_LAST_ROW,
                )
            )
    return targets


def build_label_index(targets: list[TargetRow]) -> dict[str, list[TargetRow]]:
    index: dict[str, list[TargetRow]] = {}
    for t in targets:
        index.setdefault(normalize(t.label), []).append(t)
    return index


def is_estimate_cell(cell) -> bool:
    color = cell.font.color.rgb if cell.font.color else None
    return isinstance(color, str) and color.upper() == ESTIMATE_FONT_COLOR


def _set_font_color(cell, rgb: str) -> None:
    new_font = copy(cell.font)
    new_font.color = openpyxl.styles.colors.Color(rgb=rgb)
    cell.font = new_font


def apply_month_amounts(ws, month: int, amounts_by_row: dict[int, float]) -> tuple[int, int]:
    """Estimate cells (light blue font) are replaced by the real amount and
    turned black. Cells that already hold a real value (black font) are
    added to, so incremental syncs accumulate instead of erasing what a
    previous run already wrote. Returns (nb_replaced, nb_added)."""
    col_letter = MONTH_COLUMNS[month]
    replaced = added = 0
    for row, amount in amounts_by_row.items():
        cell = ws[f"{col_letter}{row}"]
        if is_estimate_cell(cell):
            cell.value = round(amount, 2)
            _set_font_color(cell, REAL_FONT_COLOR)
            replaced += 1
        else:
            current = cell.value if isinstance(cell.value, (int, float)) else 0
            cell.value = round(current + amount, 2)
            added += 1
    return replaced, added


def write_unmatched(wb, entries: list[dict]) -> None:
    if UNMATCHED_SHEET in wb.sheetnames:
        ws = wb[UNMATCHED_SHEET]
    else:
        ws = wb.create_sheet(UNMATCHED_SHEET)
        ws.append(["Date", "Libellé", "Montant", "Tiers résolu", "Catégorie Pennylane", "Raison"])
    for entry in entries:
        ws.append(
            [
                entry["date"],
                entry["label"],
                entry["amount"],
                entry.get("tiers", ""),
                entry.get("category", ""),
                entry.get("reason", ""),
            ]
        )
